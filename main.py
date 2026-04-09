#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
macOS 桌面小工具：输入 URL，点击确定后调用第三方接口（接口逻辑在 call_third_party_api 中实现）。

说明：系统自带的 /usr/bin/python3 所带的 tkinter 依赖 Tcl/Tk，在部分 macOS 12.6 上会报错
「macOS 12 (1207) or later required」并崩溃。本程序使用 PySide6（Qt），不依赖系统 Tk。
"""

from __future__ import annotations

import csv
import json
import os
import sys
import time
import urllib.error
import urllib.request
from typing import Optional
from urllib.parse import urlparse

DETECT_TELECOM_FRAUD_URL = "http://103.90.155.233:9024/detect_telecom_fraud"
QUERY_TELECOM_FRAUD_RESULT_URL = "http://103.90.155.233:9024/query_telecom_fraud_result"

# 检测接口回调地址（与接口约定一致）
DETECT_CALLBACK_URL = (
    "https://landcall.langma.cn/notifications/voice-quality-test-result/lm-model"
)

DETECT_INDUSTRIES = {
    "金融业": [
        "银行贷款",
        "互联网贷款",
        "信用卡",
        "pos机（刷卡机）",
        "保险类",
        "投资理财",
        "股票、证券类",
        "其他",
        "债务优化",
        "互联网支付",
        "收款码",
    ],
    "专业服务": ["社保/医保/政府代办", "公检法代办"],
    "汽车": ["汽车销售", "汽车服务", "汽车配件", "汽车保险", "ETC", "其他", "饲料销售", "二手买卖"],
    "教育培训": ["学校教育", "培训/教育辅导", "其他", "消防培训", "艺术培训", "学历提升"],
    "广告/传媒/文化": ["广告/推广", "会议邀约/营销", "新闻/出版", "互联网推广", "文化传媒", "其他"],
    "零售/贸易/批发": [
        "烟酒类",
        "服装/纺织",
        "珠宝/首饰",
        "家具/家居/家电",
        "食品/饮料",
        "玩具/礼品",
        "零售",
        "批发",
        "采购",
        "其他",
        "生活用品",
        "图片批发",
        "图书",
    ],
    "交通运输/仓储/物流": ["快递", "货运/物流", "仓储", "其他"],
    "生活服务": [
        "婚纱摄影",
        "美业",
        "健身",
        "婚恋/交友",
        "机票",
        "家政",
        "旅游",
        "餐饮",
        "青少年活动中心（少年宫）",
        "其他",
        "回收",
        "保健服务",
        "酒店管理",
        "养生",
        "招生策划",
        "环保业",
        "化妆品销售",
        "会销",
        "婚礼策划",
        "书籍",
        "销售",
        "健康咨询",
    ],
    "医疗保健": ["医疗检测", ">药品", "保健品", "医疗器械", "医疗美容", "其他", "医疗用品", "医疗售后回访"],
    "互联网/通信": [
        "电商(淘宝/抖音/苏宁/京东等)",
        "电商代运营",
        "软件服务",
        "通讯业务",
        "游戏",
        "其他",
        "地图服务",
        "线上运营",
        "服务器 计算机",
        "云服务器销售",
        "地图标注",
    ],
    "房地产/建筑": ["房地产中介/销售", "物业服务", "装饰装修", "建材", "其他", "建筑服务", "物业维修"],
    "博彩/赌博": ["六合彩资料推销", "赌博引流(杀猪盘前兆)", "博彩预测/内幕消息"],
    "制造业": [
        "机械设>备",
        "电子设备",
        "农副产品",
        "渔/牧/木业",
        "家具制造",
        "化工原料/化学制品",
        "印刷/包装/造纸",
        "金属制品",
        "其他",
        "塑料加工业",
        "农业种植",
        "钢材",
        "石材",
        "包装袋",
        "加工",
    ],
    "其他": ["养老/社会保障", "环保/能源", "其他", "测试", "翡翠销售", "暂无行业", "奢侈品回收"],
}

try:
    from PySide6.QtCore import QObject, Qt, QThread, Signal
    from PySide6.QtWidgets import (
        QApplication,
        QFileDialog,
        QHBoxLayout,
        QLabel,
        QLineEdit,
        QMessageBox,
        QProgressBar,
        QPushButton,
        QTabBar,
        QVBoxLayout,
        QWidget,
    )
except ImportError:
    exe = sys.executable or "python3"
    print(
        "未安装 PySide6。请先激活虚拟环境，再安装依赖（与运行本脚本使用同一个 python）：\n"
        f'  "{exe}" -m pip install -r requirements.txt\n'
        "若未使用 venv 且提示权限不足，可改为：\n"
        f'  "{exe}" -m pip install --user -r requirements.txt\n',
        file=sys.stderr,
    )
    sys.exit(1)


def post_detect_telecom_fraud(
    audio_url: str, timeout_sec: float = 300.0, max_retries: int = 2
) -> str:
    """
    调用检测接口，返回响应中的 session_id。
    入参：audio_data_base64、audio_url、audio_codec、risk_keywords、callback_url。
    其中 audio_codec 按当前约定固定为空字符串。
    """
    body = {
        "audio_data_base64": "",
        "audio_url": audio_url,
        "audio_codec": "",
        "risk_keywords": [],
        "industries": DETECT_INDUSTRIES,
        "callback_url": DETECT_CALLBACK_URL,
    }
    payload = json.dumps(body, ensure_ascii=False).encode("utf-8")
    raw = ""
    for attempt in range(max_retries + 1):
        req = urllib.request.Request(
            DETECT_TELECOM_FRAUD_URL,
            data=payload,
            method="POST",
            headers={"Content-Type": "application/json; charset=utf-8"},
        )
        try:
            with urllib.request.urlopen(req, timeout=timeout_sec) as resp:
                raw = resp.read().decode("utf-8")
                break
        except urllib.error.HTTPError as e:
            err_body = e.read().decode("utf-8", errors="replace")
            if e.code == 503 and attempt < max_retries:
                # 服务临时不可用时做短暂重试
                time.sleep(1.5 * (attempt + 1))
                continue
            if e.code == 503:
                raise RuntimeError("HTTP 503：Service Unavailable（服务繁忙，请稍后重试）") from e
            raise RuntimeError(f"HTTP {e.code}：{err_body or e.reason}") from e
        except urllib.error.URLError as e:
            raise RuntimeError(f"网络错误：{e.reason}") from e

    try:
        data = json.loads(raw)
    except json.JSONDecodeError as e:
        raise ValueError(f"响应不是合法 JSON：{raw[:500]}") from e

    if not isinstance(data, dict):
        raise ValueError(f"响应格式异常（应为 JSON 对象）：{raw[:500]}")
    if "session_id" not in data:
        raise ValueError(f"响应中缺少 session_id：{raw[:500]}")
    return str(data["session_id"])


def post_query_telecom_fraud_result(session_id: str, timeout_sec: float = 300.0) -> dict:
    """
    调用查询接口，返回包含 reason/risk_confidence/risk_severity/exist_risk 的结果。
    """
    # 接口实际校验字段为 session_ids（数组）
    payload = json.dumps({"session_ids": [session_id]}, ensure_ascii=False).encode("utf-8")
    req = urllib.request.Request(
        QUERY_TELECOM_FRAUD_RESULT_URL,
        data=payload,
        method="POST",
        headers={"Content-Type": "application/json; charset=utf-8"},
    )
    try:
        with urllib.request.urlopen(req, timeout=timeout_sec) as resp:
            raw = resp.read().decode("utf-8")
    except urllib.error.HTTPError as e:
        err_body = e.read().decode("utf-8", errors="replace")
        raise RuntimeError(f"HTTP {e.code}：{err_body or e.reason}") from e
    except urllib.error.URLError as e:
        raise RuntimeError(f"网络错误：{e.reason}") from e

    try:
        data = json.loads(raw)
    except json.JSONDecodeError as e:
        raise ValueError(f"响应不是合法 JSON：{raw[:500]}") from e

    # 兼容三类响应：
    # 1) 顶层直接是对象，且包含 reason/risk_confidence/...
    # 2) 顶层是对象，但结果在 results/data 的列表里
    # 3) 顶层直接是列表，取第一项
    if isinstance(data, list):
        if data and isinstance(data[0], dict):
            return data[0]
        raise ValueError(f"响应格式异常（列表为空或元素非对象）：{raw[:500]}")

    if not isinstance(data, dict):
        raise ValueError(f"响应格式异常（应为 JSON 对象或对象列表）：{raw[:500]}")

    if "reason" in data or "risk_confidence" in data or "risk_severity" in data or "exist_risk" in data:
        return data

    results = data.get("results")
    if isinstance(results, list) and results and isinstance(results[0], dict):
        return results[0]

    items = data.get("data")
    if isinstance(items, list) and items and isinstance(items[0], dict):
        return items[0]

    raise ValueError(f"响应中未找到风险字段：{raw[:500]}")


def read_urls_from_single_column_file(file_path: str) -> list[str]:
    """
    从单列表格读取 URL。
    支持 .csv/.txt/.xlsx（.txt 为每行一个 URL，其余读取第一列）。
    """
    ext = os.path.splitext(file_path)[1].lower()
    urls: list[str] = []

    if ext == ".txt":
        with open(file_path, "r", encoding="utf-8") as f:
            for line in f:
                val = line.strip()
                if val:
                    urls.append(val)
    elif ext == ".csv":
        with open(file_path, "r", encoding="utf-8-sig", newline="") as f:
            reader = csv.reader(f)
            for row in reader:
                if not row:
                    continue
                val = str(row[0]).strip()
                # 跳过常见表头
                if val.lower() in {"url", "audio_url"}:
                    continue
                if val:
                    urls.append(val)
    elif ext == ".xlsx":
        try:
            from openpyxl import load_workbook  # type: ignore[reportMissingImports]
        except ImportError as e:
            raise ValueError(
                "缺少 openpyxl 依赖，无法读取 .xlsx。请先执行：pip install -r requirements.txt"
            ) from e

        wb = load_workbook(file_path, read_only=True, data_only=True)
        try:
            ws = wb.active
            for row in ws.iter_rows(min_col=1, max_col=1, values_only=True):
                cell = row[0]
                if cell is None:
                    continue
                val = str(cell).strip()
                if val.lower() in {"url", "audio_url"}:
                    continue
                if val:
                    urls.append(val)
        finally:
            wb.close()
    else:
        raise ValueError("仅支持 .csv/.txt/.xlsx 文件（单列 URL）。")

    if not urls:
        raise ValueError("文件中未读取到 URL，请检查是否为单列且有内容。")
    return urls


def export_url_session_map(rows: list[dict], save_path: str) -> None:
    """
    导出批量结果到本地 CSV。
    """
    with open(save_path, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=["url", "session_id", "error"])
        writer.writeheader()
        for row in rows:
            writer.writerow(
                {
                    "url": row.get("url", ""),
                    "session_id": row.get("session_id", ""),
                    "error": row.get("error", ""),
                }
            )


def read_url_session_rows_from_file(file_path: str) -> list[dict]:
    """
    从表格读取 url + session_id 两列数据。
    要求表头包含：url, session_id（大小写不敏感）。
    支持 .csv/.xlsx。
    """
    ext = os.path.splitext(file_path)[1].lower()
    rows: list[dict] = []

    if ext == ".csv":
        with open(file_path, "r", encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f)
            if not reader.fieldnames:
                raise ValueError("CSV 缺少表头，需包含 url 和 session_id。")
            normalized = {str(n).strip().lower(): n for n in reader.fieldnames if n is not None}
            if "url" not in normalized or "session_id" not in normalized:
                raise ValueError("CSV 表头需包含 url 和 session_id。")
            url_key = normalized["url"]
            sid_key = normalized["session_id"]
            for item in reader:
                url = str(item.get(url_key, "")).strip()
                session_id = str(item.get(sid_key, "")).strip()
                if not url and not session_id:
                    continue
                rows.append({"url": url, "session_id": session_id})
    elif ext == ".xlsx":
        try:
            from openpyxl import load_workbook  # type: ignore[reportMissingImports]
        except ImportError as e:
            raise ValueError(
                "缺少 openpyxl 依赖，无法读取 .xlsx。请先执行：pip install -r requirements.txt"
            ) from e

        wb = load_workbook(file_path, read_only=True, data_only=True)
        try:
            ws = wb.active
            header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
            if not header:
                raise ValueError("Excel 缺少表头，需包含 url 和 session_id。")
            header_norm = [str(h).strip().lower() if h is not None else "" for h in header]
            if "url" not in header_norm or "session_id" not in header_norm:
                raise ValueError("Excel 表头需包含 url 和 session_id。")
            url_idx = header_norm.index("url")
            sid_idx = header_norm.index("session_id")
            max_col = max(url_idx, sid_idx) + 1
            for row in ws.iter_rows(min_row=2, max_col=max_col, values_only=True):
                url = str(row[url_idx]).strip() if url_idx < len(row) and row[url_idx] is not None else ""
                session_id = (
                    str(row[sid_idx]).strip() if sid_idx < len(row) and row[sid_idx] is not None else ""
                )
                if not url and not session_id:
                    continue
                rows.append({"url": url, "session_id": session_id})
        finally:
            wb.close()
    else:
        raise ValueError("批量查询仅支持 .csv/.xlsx 文件（需含 url、session_id 表头）。")

    if not rows:
        raise ValueError("文件中未读取到有效数据，请检查 url 与 session_id 列内容。")
    return rows


def export_batch_query_rows(rows: list[dict], save_path: str) -> None:
    """
    导出批量查询结果到本地 CSV。
    """
    fieldnames = [
        "url",
        "session_id",
        "industry",
        "reason",
        "risk_confidence",
        "risk_severity",
        "exist_risk",
        "asr_text",
        "error",
    ]
    with open(save_path, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow({k: row.get(k, "") for k in fieldnames})


class DetectWorker(QObject):
    """在子线程中请求接口，避免阻塞界面。"""

    finished_ok = Signal(str)
    finished_err = Signal(str)

    def __init__(self, audio_url: str) -> None:
        super().__init__()
        self._audio_url = audio_url

    def run(self) -> None:
        try:
            session_id = post_detect_telecom_fraud(self._audio_url)
            self.finished_ok.emit(session_id)
        except Exception as e:
            self.finished_err.emit(str(e))


class QueryWorker(QObject):
    """在子线程中查询 session_id 对应结果，避免阻塞界面。"""

    finished_ok = Signal(dict)
    finished_err = Signal(str)

    def __init__(self, session_id: str) -> None:
        super().__init__()
        self._session_id = session_id

    def run(self) -> None:
        try:
            result = post_query_telecom_fraud_result(self._session_id)
            self.finished_ok.emit(result)
        except Exception as e:
            self.finished_err.emit(str(e))


class BatchDetectWorker(QObject):
    """批量提交 URL 并返回 url-session_id 对照结果。"""

    finished_ok = Signal(list)
    finished_err = Signal(str)
    progress_changed = Signal(int, int)

    def __init__(self, urls: list[str]) -> None:
        super().__init__()
        self._urls = urls

    def run(self) -> None:
        rows: list[dict] = []
        try:
            total = len(self._urls)
            for idx, url in enumerate(self._urls, start=1):
                if not is_valid_http_url(url):
                    rows.append({"url": url, "session_id": "", "error": "URL 格式无效"})
                    self.progress_changed.emit(idx, total)
                    continue
                try:
                    session_id = post_detect_telecom_fraud(url)
                    rows.append({"url": url, "session_id": session_id, "error": ""})
                except Exception as e:
                    rows.append({"url": url, "session_id": "", "error": str(e)})
                self.progress_changed.emit(idx, total)
            self.finished_ok.emit(rows)
        except Exception as e:
            self.finished_err.emit(str(e))


class BatchQueryWorker(QObject):
    """批量查询 session_id 对应结果。"""

    finished_ok = Signal(list)
    finished_err = Signal(str)
    progress_changed = Signal(int, int)

    def __init__(self, items: list[dict]) -> None:
        super().__init__()
        self._items = items

    def run(self) -> None:
        rows: list[dict] = []
        try:
            total = len(self._items)
            for idx, item in enumerate(self._items, start=1):
                url = str(item.get("url", "")).strip()
                session_id = str(item.get("session_id", "")).strip()
                row = {
                    "url": url,
                    "session_id": session_id,
                    "industry": "",
                    "reason": "",
                    "risk_confidence": "",
                    "risk_severity": "",
                    "exist_risk": "",
                    "asr_text": "",
                    "error": "",
                }
                if not session_id:
                    row["error"] = "session_id 为空"
                    rows.append(row)
                    self.progress_changed.emit(idx, total)
                    continue
                try:
                    result = post_query_telecom_fraud_result(session_id)
                    row["industry"] = result.get("industry", "")
                    row["reason"] = result.get("reason", "")
                    row["risk_confidence"] = result.get("risk_confidence", "")
                    row["risk_severity"] = result.get("risk_severity", "")
                    row["exist_risk"] = result.get("exist_risk", "")
                    row["asr_text"] = result.get("asr_text", "")
                except Exception as e:
                    row["error"] = str(e)
                rows.append(row)
                self.progress_changed.emit(idx, total)
            self.finished_ok.emit(rows)
        except Exception as e:
            self.finished_err.emit(str(e))


def is_valid_http_url(url: str) -> bool:
    url = url.strip()
    if not url:
        return False
    try:
        parsed = urlparse(url)
    except Exception:
        return False
    return parsed.scheme in ("http", "https") and bool(parsed.netloc)


class MainWindow(QWidget):
    def __init__(self) -> None:
        super().__init__()
        self.setWindowTitle("URL 工具")
        self.setMinimumWidth(480)
        # 整体高度收紧约三分之一，减少纵向空白
        self.setMinimumHeight(410)

        self._thread: Optional[QThread] = None
        self._worker: Optional[QObject] = None
        self._batch_rows: Optional[list[dict]] = None

        layout = QVBoxLayout(self)
        layout.setSpacing(8)

        self.mode_tabs = QTabBar()
        self.mode_tabs.addTab("单个查询")
        self.mode_tabs.addTab("文件上传")
        self.mode_tabs.currentChanged.connect(self._on_tab_changed)
        layout.addWidget(self.mode_tabs)

        self.url_label = QLabel("请输入 URL：")
        layout.addWidget(self.url_label)

        self.url_edit = QLineEdit()
        self.url_edit.setPlaceholderText("https://example.com")
        self.url_edit.returnPressed.connect(self.on_query_auto)
        layout.addWidget(self.url_edit)

        self.session_id_label = QLabel("请输入 session_id：")
        layout.addWidget(self.session_id_label)
        self.session_id_edit = QLineEdit()
        self.session_id_edit.setPlaceholderText("请输入接口返回的 session_id")
        self.session_id_edit.returnPressed.connect(self.on_query_auto)
        layout.addWidget(self.session_id_edit)

        self.detect_section_label = QLabel("批量提交（单列 URL，支持 .csv/.txt/.xlsx）：")
        layout.addWidget(self.detect_section_label)
        detect_file_row = QHBoxLayout()
        self.batch_detect_file_edit = QLineEdit()
        self.batch_detect_file_edit.setPlaceholderText("请选择用于批量提交的 URL 文件")
        self.batch_detect_file_edit.setReadOnly(True)
        self.btn_pick_detect_file = QPushButton("选择提交文件")
        self.btn_pick_detect_file.clicked.connect(self.on_pick_detect_file)
        detect_file_row.addWidget(self.batch_detect_file_edit)
        detect_file_row.addWidget(self.btn_pick_detect_file)
        self.detect_file_row_widgets = [self.batch_detect_file_edit, self.btn_pick_detect_file]
        layout.addLayout(detect_file_row)

        self.batch_detect_progress_label = QLabel("批量提交进度：0%")
        self.batch_detect_progress_bar = QProgressBar()
        self.batch_detect_progress_bar.setRange(0, 100)
        self.batch_detect_progress_bar.setValue(0)
        layout.addWidget(self.batch_detect_progress_label)
        layout.addWidget(self.batch_detect_progress_bar)

        self.query_section_label = QLabel("批量查询（表头需包含 url、session_id，支持 .csv/.xlsx）：")
        layout.addWidget(self.query_section_label)
        query_file_row = QHBoxLayout()
        self.batch_query_file_edit = QLineEdit()
        self.batch_query_file_edit.setPlaceholderText("请选择用于批量查询的文件")
        self.batch_query_file_edit.setReadOnly(True)
        self.btn_pick_query_file = QPushButton("选择查询文件")
        self.btn_pick_query_file.clicked.connect(self.on_pick_query_file)
        query_file_row.addWidget(self.batch_query_file_edit)
        query_file_row.addWidget(self.btn_pick_query_file)
        self.query_file_row_widgets = [self.batch_query_file_edit, self.btn_pick_query_file]
        layout.addLayout(query_file_row)

        self.batch_query_progress_label = QLabel("批量查询进度：0%")
        self.batch_query_progress_bar = QProgressBar()
        self.batch_query_progress_bar.setRange(0, 100)
        self.batch_query_progress_bar.setValue(0)
        layout.addWidget(self.batch_query_progress_label)
        layout.addWidget(self.batch_query_progress_bar)

        row = QHBoxLayout()
        row.addStretch()
        btn_quit = QPushButton("退出")
        btn_quit.clicked.connect(self.close)
        self.btn_reset = QPushButton("重置")
        self.btn_reset.clicked.connect(self.on_reset)
        self.btn_query = QPushButton("查询")
        self.btn_query.setDefault(True)
        self.btn_query.clicked.connect(self.on_query_auto)
        row.addWidget(btn_quit)
        row.addWidget(self.btn_reset)
        row.addWidget(self.btn_query)
        layout.addLayout(row)

        self.url_edit.setFocus(Qt.FocusReason.OtherFocusReason)
        self._on_tab_changed(0)

    def _set_busy(self, busy: bool) -> None:
        self.btn_query.setEnabled(not busy)
        self.btn_reset.setEnabled(not busy)
        self.btn_pick_detect_file.setEnabled(not busy)
        self.btn_pick_query_file.setEnabled(not busy)
        self.mode_tabs.setEnabled(not busy)
        self.url_edit.setEnabled(not busy)
        self.session_id_edit.setEnabled(not busy)

    def _set_widgets_visible(self, widgets: list[QWidget], visible: bool) -> None:
        for w in widgets:
            w.setVisible(visible)

    def _on_tab_changed(self, index: int) -> None:
        is_single_mode = index == 0
        is_file_mode = not is_single_mode

        self.url_label.setVisible(is_single_mode)
        self.url_edit.setVisible(is_single_mode)
        self.session_id_label.setVisible(is_single_mode)
        self.session_id_edit.setVisible(is_single_mode)

        self.detect_section_label.setVisible(is_file_mode)
        self.query_section_label.setVisible(is_file_mode)
        self._set_widgets_visible(self.detect_file_row_widgets, is_file_mode)
        self._set_widgets_visible(self.query_file_row_widgets, is_file_mode)
        self.batch_detect_progress_label.setVisible(is_file_mode)
        self.batch_detect_progress_bar.setVisible(is_file_mode)
        self.batch_query_progress_label.setVisible(is_file_mode)
        self.batch_query_progress_bar.setVisible(is_file_mode)

        self.btn_query.setText("查询" if is_single_mode else "执行")
        self.btn_reset.setText("重置")

    def _reset_detect_progress(self) -> None:
        self.batch_detect_progress_bar.setValue(0)
        self.batch_detect_progress_label.setText("批量提交进度：0%")

    def _reset_query_progress(self) -> None:
        self.batch_query_progress_bar.setValue(0)
        self.batch_query_progress_label.setText("批量查询进度：0%")

    def _on_batch_detect_progress(self, current: int, total: int) -> None:
        percent = int((current / total) * 100) if total else 0
        self.batch_detect_progress_bar.setValue(percent)
        self.batch_detect_progress_label.setText(f"批量提交进度：{percent}%（{current}/{total}）")

    def _on_batch_query_progress(self, current: int, total: int) -> None:
        percent = int((current / total) * 100) if total else 0
        self.batch_query_progress_bar.setValue(percent)
        self.batch_query_progress_label.setText(f"批量查询进度：{percent}%（{current}/{total}）")

    def _on_detect_ok(self, session_id: str) -> None:
        self._set_busy(False)
        self._cleanup_thread()
        self.session_id_edit.setText(session_id)
        QMessageBox.information(self, "session_id", f"session_id：\n{session_id}")

    def _on_detect_err(self, message: str) -> None:
        self._set_busy(False)
        self._cleanup_thread()
        QMessageBox.critical(self, "接口失败", message)

    def _on_batch_ok(self, rows: list[dict]) -> None:
        self._set_busy(False)
        self._cleanup_thread()
        self._batch_rows = rows
        self.batch_detect_progress_bar.setValue(100)

        default_name = "url_session_id结果.csv"
        save_path, _ = QFileDialog.getSaveFileName(
            self,
            "保存批量结果",
            default_name,
            "CSV Files (*.csv)",
        )
        if not save_path:
            QMessageBox.information(self, "提示", "已完成批量请求，但你取消了保存。")
            return
        if not save_path.lower().endswith(".csv"):
            save_path += ".csv"

        try:
            export_url_session_map(rows, save_path)
        except Exception as e:
            QMessageBox.critical(self, "保存失败", f"结果已生成，但保存文件失败：{e}")
            return

        success_count = sum(1 for r in rows if r.get("session_id"))
        fail_count = len(rows) - success_count
        QMessageBox.information(
            self,
            "批量完成",
            f"共处理 {len(rows)} 条。\n成功 {success_count} 条，失败 {fail_count} 条。\n已保存到：\n{save_path}",
        )

    def _on_batch_query_ok(self, rows: list[dict]) -> None:
        self._set_busy(False)
        self._cleanup_thread()
        self.batch_query_progress_bar.setValue(100)

        default_name = "session_id查询结果.csv"
        save_path, _ = QFileDialog.getSaveFileName(
            self,
            "保存批量查询结果",
            default_name,
            "CSV Files (*.csv)",
        )
        if not save_path:
            QMessageBox.information(self, "提示", "已完成批量查询，但你取消了保存。")
            return
        if not save_path.lower().endswith(".csv"):
            save_path += ".csv"

        try:
            export_batch_query_rows(rows, save_path)
        except Exception as e:
            QMessageBox.critical(self, "保存失败", f"结果已生成，但保存文件失败：{e}")
            return

        success_count = sum(1 for r in rows if not r.get("error"))
        fail_count = len(rows) - success_count
        QMessageBox.information(
            self,
            "批量查询完成",
            f"共处理 {len(rows)} 条。\n成功 {success_count} 条，失败 {fail_count} 条。\n已保存到：\n{save_path}",
        )

    def _cleanup_thread(self) -> None:
        if self._thread is not None:
            self._thread.quit()
            self._thread.wait()
        self._thread = None
        self._worker = None

    def wait_for_thread_before_exit(self) -> None:
        """
        应用退出兜底：确保后台线程已结束，避免 QThread 被提前销毁。
        """
        if self._thread is not None and self._thread.isRunning():
            self._thread.wait()
            self._thread = None
            self._worker = None

    def closeEvent(self, event) -> None:
        # 防止线程未结束时窗口直接销毁导致：
        # QThread: Destroyed while thread is still running
        if self._thread is not None and self._thread.isRunning():
            QMessageBox.warning(self, "请稍候", "后台任务仍在执行，请等待完成后再退出。")
            event.ignore()
            return
        event.accept()

    def on_confirm(self) -> None:
        if self._thread is not None:
            return
        url = self.url_edit.text().strip()
        if not is_valid_http_url(url):
            QMessageBox.warning(
                self,
                "输入无效",
                "请输入以 http:// 或 https:// 开头的完整 URL。",
            )
            return

        self._set_busy(True)
        self._thread = QThread()
        self._worker = DetectWorker(url)
        self._worker.moveToThread(self._thread)
        self._thread.started.connect(self._worker.run)
        self._worker.finished_ok.connect(self._on_detect_ok)
        self._worker.finished_err.connect(self._on_detect_err)
        self._thread.start()

    def _on_query_ok(self, result: dict) -> None:
        self._set_busy(False)
        self._cleanup_thread()
        message = (
            f"industry: {result.get('industry', '')}\n"
            f"reason: {result.get('reason', '')}\n"
            f"risk_confidence: {result.get('risk_confidence', '')}\n"
            f"risk_severity: {result.get('risk_severity', '')}\n"
            f"exist_risk: {result.get('exist_risk', '')}\n"
            f"asr_text: {result.get('asr_text', '')}"
        )
        QMessageBox.information(self, "查询结果", message)

    def on_query_confirm(self) -> None:
        if self._thread is not None:
            return
        session_id = self.session_id_edit.text().strip()
        if not session_id:
            QMessageBox.warning(self, "输入无效", "请输入 session_id。")
            return

        self._set_busy(True)
        self._thread = QThread()
        self._worker = QueryWorker(session_id)
        self._worker.moveToThread(self._thread)
        self._thread.started.connect(self._worker.run)
        self._worker.finished_ok.connect(self._on_query_ok)
        self._worker.finished_err.connect(self._on_detect_err)
        self._thread.start()

    def on_pick_detect_file(self) -> None:
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "选择单列 URL 文件",
            "",
            "URL Files (*.csv *.txt *.xlsx)",
        )
        if file_path:
            self.batch_detect_file_edit.setText(file_path)

    def on_pick_query_file(self) -> None:
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "选择批量查询文件",
            "",
            "Query Files (*.csv *.xlsx)",
        )
        if file_path:
            self.batch_query_file_edit.setText(file_path)

    def on_batch_confirm(self) -> None:
        if self.mode_tabs.currentIndex() != 1:
            QMessageBox.information(self, "提示", "请切换到“文件上传”选项卡后再执行批量提交。")
            return
        if self._thread is not None:
            return
        file_path = self.batch_detect_file_edit.text().strip()
        if not file_path:
            QMessageBox.warning(self, "输入无效", "请先选择 URL 文件。")
            return

        try:
            urls = read_urls_from_single_column_file(file_path)
        except Exception as e:
            QMessageBox.critical(self, "文件读取失败", str(e))
            return

        self._set_busy(True)
        self._reset_detect_progress()
        self._thread = QThread()
        self._worker = BatchDetectWorker(urls)
        self._worker.moveToThread(self._thread)
        self._thread.started.connect(self._worker.run)
        self._worker.progress_changed.connect(self._on_batch_detect_progress)
        self._worker.finished_ok.connect(self._on_batch_ok)
        self._worker.finished_err.connect(self._on_detect_err)
        self._thread.start()

    def on_batch_query_confirm(self) -> None:
        if self.mode_tabs.currentIndex() != 1:
            QMessageBox.information(self, "提示", "请切换到“文件上传”选项卡后再执行批量查询。")
            return
        if self._thread is not None:
            return
        file_path = self.batch_query_file_edit.text().strip()
        if not file_path:
            QMessageBox.warning(self, "输入无效", "请先选择批量查询文件。")
            return

        try:
            items = read_url_session_rows_from_file(file_path)
        except Exception as e:
            QMessageBox.critical(self, "文件读取失败", str(e))
            return

        self._set_busy(True)
        self._reset_query_progress()
        self._thread = QThread()
        self._worker = BatchQueryWorker(items)
        self._worker.moveToThread(self._thread)
        self._thread.started.connect(self._worker.run)
        self._worker.progress_changed.connect(self._on_batch_query_progress)
        self._worker.finished_ok.connect(self._on_batch_query_ok)
        self._worker.finished_err.connect(self._on_detect_err)
        self._thread.start()

    def on_query_auto(self) -> None:
        """
        自动判断查询模式：
        - 单个查询页：执行单条查询；
        - 文件上传页：自动判断执行批量提交或批量查询。
        """
        if self.mode_tabs.currentIndex() == 1:
            self.on_file_auto()
            return
        url = self.url_edit.text().strip()
        session_id = self.session_id_edit.text().strip()
        if url:
            self.on_confirm()
            return
        if session_id:
            self.on_query_confirm()
            return
        QMessageBox.warning(self, "输入无效", "请先输入 URL，或输入 session_id。")

    def on_file_auto(self) -> None:
        """
        文件上传页自动判断：
        - 若选择了批量查询文件，优先执行批量查询；
        - 否则若选择了批量提交文件，执行批量提交；
        - 两者都没选则提示。
        """
        query_file = self.batch_query_file_edit.text().strip()
        detect_file = self.batch_detect_file_edit.text().strip()
        if query_file:
            self.on_batch_query_confirm()
            return
        if detect_file:
            self.on_batch_confirm()
            return
        QMessageBox.warning(self, "输入无效", "请先选择批量提交文件或批量查询文件。")

    def on_reset(self) -> None:
        """
        根据当前选项卡重置输入与进度，方便二次查询。
        """
        if self._thread is not None:
            QMessageBox.information(self, "提示", "后台任务执行中，请完成后再重置。")
            return

        if self.mode_tabs.currentIndex() == 0:
            self.url_edit.clear()
            self.session_id_edit.clear()
            self.url_edit.setFocus(Qt.FocusReason.OtherFocusReason)
            return

        self.batch_detect_file_edit.clear()
        self.batch_query_file_edit.clear()
        self._reset_detect_progress()
        self._reset_query_progress()


def main() -> None:
    app = QApplication([])
    w = MainWindow()
    app.aboutToQuit.connect(w.wait_for_thread_before_exit)
    w.show()
    app.exec()


if __name__ == "__main__":
    main()
